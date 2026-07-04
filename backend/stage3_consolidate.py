"""
Stage 3: Incremental Master Template Consolidation.

Supported phases:
- early: template + CAT1 + ASS1 -> output
- mid: existing output + CAT2 + ASS2 -> output
- terminal: existing output + Terminal -> output
- end: no new consolidation; validate output exists
"""

import json
import os
import sys
import re
from openpyxl import load_workbook
from utils import validate_file_exists, normalize_question_id, clean_numeric


def _clean_co_tag(value):
    if value is None:
        return ""
    s = str(value).strip()
    match = re.search(r"CO\d+", s, re.IGNORECASE)
    if match:
        return match.group(0).upper()
    return ""

def _to_numeric(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip().replace("'", "")
    if not s or s.lower() in ("nan", "none", ""):
        return None
    try:
        if re.fullmatch(r"-?\d+", s):
            return int(s)
        if re.fullmatch(r"-?\d*\.\d+", s):
            return float(s)
    except (TypeError, ValueError):
        pass
    return None

def get_main_question(q_label):
    q = str(q_label).strip()
    if "." in q:
        return q.split(".")[0]
    return q


def build_dynamic_question_mapping(question_data):
    parts = ["A", "B", "C", "D", "E", "F"]
    part_idx = 0
    current_part = parts[part_idx]
    
    prev_co = 0
    part_q_count = 0
    
    mapping = {}
    main_q_assigned = {}
    
    for qd in question_data:
        q_label = qd['q_label']
        co_num = qd['co_num']
        
        main_q = get_main_question(q_label)
        
        if main_q not in main_q_assigned:
            if co_num > 0 and prev_co > 0 and co_num < prev_co:
                part_idx += 1
                if part_idx < len(parts):
                    current_part = parts[part_idx]
                part_q_count = 0
            
            part_q_count += 1
            template_id = f"{current_part}{part_q_count}"
            main_q_assigned[main_q] = template_id
            if co_num > 0:
                prev_co = co_num
                
        mapping[q_label] = main_q_assigned[main_q]
        
    return mapping


def process_cat(cat_file, template_ws, template_start_col, template_end_col):
    """
    Process a RAW CAMU CAT file and merge into template using dynamic mapping.
    """
    cat_wb = load_workbook(cat_file, data_only=True)
    cat_ws = cat_wb.active

    # Find first row where column A has a valid Register Number (digits)
    student_start = 5  # default fallback
    for r in range(3, cat_ws.max_row + 1):
        val = cat_ws.cell(row=r, column=1).value
        if val is not None and re.match(r"^\d+", str(val).strip()):
            student_start = r
            break

    camu_row_q_labels  = 1
    camu_row_max_marks = student_start - 1
    camu_row_co_tags   = student_start - 2
    camu_marks_start_col = 7  # Column G (1-indexed)

    # Read template questions
    template_q = {}  # {normalized_q: col}
    for col in range(template_start_col, template_end_col + 1):
        q = normalize_question_id(template_ws.cell(row=6, column=col).value)
        if q and q != "TOTAL":
            template_q[q] = col

    # 1. Read all question columns
    question_data = []
    col = camu_marks_start_col
    while col <= cat_ws.max_column:
        q_label  = cat_ws.cell(row=camu_row_q_labels,  column=col).value
        co_raw   = cat_ws.cell(row=camu_row_co_tags,   column=col).value
        max_raw  = cat_ws.cell(row=camu_row_max_marks,  column=col).value

        if q_label is None and co_raw is None and max_raw is None:
            break

        co_str = _clean_co_tag(co_raw)
        co_num = 0
        if co_str:
            match = re.search(r"CO(\d+)", co_str, re.IGNORECASE)
            if match:
                co_num = int(match.group(1))

        if q_label is None:
            q_label = str(col - camu_marks_start_col + 1)
        else:
            q_label = str(q_label).strip()

        max_m = _to_numeric(max_raw)

        question_data.append({
            'q_label': q_label,
            'co_str': co_str,
            'co_num': co_num,
            'max_m': max_m,
            'col': col
        })
        col += 1

    # 2. Build mapping
    mapping = build_dynamic_question_mapping(question_data)
    
    # 3. Combine max marks & COs per template column
    template_col_info = {} # tcol -> {'co': set(), 'max_m': val}
    for qd in question_data:
        t_id = mapping.get(qd['q_label'])
        if t_id and t_id in template_q:
            tcol = template_q[t_id]
            if tcol not in template_col_info:
                template_col_info[tcol] = {'co': set(), 'max_m': 0}
                
            if qd['co_str']:
                template_col_info[tcol]['co'].add(qd['co_str'])
            if qd['max_m']:
                if qd['max_m'] > template_col_info[tcol]['max_m']:
                    template_col_info[tcol]['max_m'] = qd['max_m']

    # Overwrite template headers
    for tcol, info in template_col_info.items():
        if info['co']:
            template_ws.cell(row=5, column=tcol).value = ", ".join(sorted(list(info['co'])))
        if info['max_m'] > 0:
            template_ws.cell(row=7, column=tcol).value = info['max_m']

    # 4. Read student data and copy to template
    row = student_start
    template_student_row = 8
    while row <= cat_ws.max_row:
        reg_val = cat_ws.cell(row=row, column=1).value
        name_val = cat_ws.cell(row=row, column=2).value
        reg_str = str(reg_val).strip() if reg_val is not None else ""
        if not reg_str or reg_str.lower() in ("nan", "none"):
            break
        if "instruction" in reg_str.lower():
            row += 1
            continue

        template_ws.cell(row=template_student_row, column=2).value = reg_str
        template_ws.cell(row=template_student_row, column=3).value = str(name_val).strip() if name_val is not None else ""

        # Aggregate marks for either/or questions
        student_marks = {} # tcol -> list of marks
        
        for qd in question_data:
            t_id = mapping.get(qd['q_label'])
            if t_id and t_id in template_q:
                tcol = template_q[t_id]
                mark = _to_numeric(cat_ws.cell(row=row, column=qd['col']).value)
                if mark is not None:
                    if tcol not in student_marks:
                        student_marks[tcol] = []
                    student_marks[tcol].append(mark)

        # Write max mark obtained
        for tcol, marks in student_marks.items():
            if marks:
                template_ws.cell(row=template_student_row, column=tcol).value = max(marks)

        row += 1
        template_student_row += 1

def process_assignment(
    ass_file, template_ws, source_start_col, template_start_col, max_cols=6, target_total=40
):
    ass_wb = load_workbook(ass_file, data_only=True)
    ass_ws = ass_wb.active

    student_start = 5
    row = student_start
    while ass_ws.cell(row=row, column=1).value is not None and str(ass_ws.cell(row=row, column=1).value).strip() != "":
        row += 1
    student_count = row - student_start

    ass_cols = 0
    while ass_ws.cell(row=3, column=source_start_col + ass_cols).value is not None and str(ass_ws.cell(row=3, column=source_start_col + ass_cols).value).strip().lower() not in ("", "none", "nan"):
        ass_cols += 1

    if ass_cols > max_cols:
        ass_cols = max_cols

    col_max_values = [0] * max_cols
    
    for c in range(ass_cols):
        co_raw = ass_ws.cell(row=3, column=source_start_col + c).value
        co_str = _clean_co_tag(co_raw)
        
        t_col_offset = c
        if co_str:
            match = re.search(r"CO(\d+)", co_str, re.IGNORECASE)
            if match:
                co_num = int(match.group(1))
                if 1 <= co_num <= max_cols:
                    t_col_offset = co_num - 1
                    
        target_col = template_start_col + t_col_offset

        values = []
        for i in range(student_count):
            val = clean_numeric(ass_ws.cell(row=student_start + i, column=source_start_col + c).value)
            template_ws.cell(row=8 + i, column=target_col).value = val
            if isinstance(val, (int, float)):
                values.append(val)

        col_max = max(values) if values else 0
        col_max_values[t_col_offset] = col_max

    current_sum = sum(col_max_values)
    diff = target_total - current_sum

    if diff != 0 and current_sum > 0:
        for idx in range(max_cols - 1, -1, -1):
            if col_max_values[idx] > 0:
                col_max_values[idx] += diff
                break

    for idx in range(max_cols):
        if col_max_values[idx] > 0:
            template_ws.cell(row=7, column=template_start_col + idx).value = col_max_values[idx]


def resolve_default_template_path():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    candidate = os.path.join(script_dir, "data", "template.xlsx")
    normalized = os.path.normpath(candidate)
    if os.path.exists(normalized):
        return normalized
    return None


def require_path(path_value, key_name):
    if not path_value:
        raise ValueError(f"Missing required path: {key_name}")
    return path_value


def load_workbook_for_phase(phase, template_path, output_path):
    if phase == "early":
        source = require_path(template_path, "template_path")
        validate_file_exists(source)
        return load_workbook(source)

    if phase == "mid" or phase == "terminal":
        source = require_path(template_path, "template_path")
        validate_file_exists(source)
        return load_workbook(source)

    if phase == "end":
        source = require_path(output_path, "output_path")
        validate_file_exists(source)
        return None

    raise ValueError("Invalid phase. Expected one of: early, mid, terminal, end")


def main():
    try:
        if len(sys.argv) < 2:
            raise ValueError("Missing JSON argument payload")

        args = json.loads(sys.argv[1])

        phase_raw = args.get("phase")
        phase = str(phase_raw).strip().lower() if phase_raw is not None else None
        output_path = require_path(args.get("output_path"), "output_path")

        template_path = args.get("template_path") or resolve_default_template_path()
        cat1_path = args.get("cat1_path")
        cat2_path = args.get("cat2_path")
        ass1_path = args.get("ass1_path")
        ass2_path = args.get("ass2_path")

        if phase == "early":
            validate_file_exists(require_path(cat1_path, "cat1_path"))
            validate_file_exists(require_path(ass1_path, "ass1_path"))

            template_wb = load_workbook_for_phase(phase, template_path, output_path)
            template_ws = template_wb.worksheets[1]

            process_cat(cat1_path, template_ws, 5, 25)
            process_assignment(ass1_path, template_ws, source_start_col=7, template_start_col=56)

            template_wb.save(output_path)

        elif phase == "mid":
            validate_file_exists(require_path(cat2_path, "cat2_path"))
            validate_file_exists(require_path(ass2_path, "ass2_path"))

            template_wb = load_workbook_for_phase(phase, template_path, output_path)
            template_ws = template_wb.worksheets[1]

            process_cat(cat2_path, template_ws, 27, 46)
            process_assignment(ass2_path, template_ws, source_start_col=7, template_start_col=64)

            template_wb.save(output_path)
            
        elif phase == "terminal":
            # Pass through the template to the output path so stage4 can use it
            template_wb = load_workbook_for_phase(phase, template_path, output_path)
            template_wb.save(output_path)

        elif phase == "end":
            load_workbook_for_phase(phase, template_path, output_path)

        else:
            raise ValueError("Invalid phase. Expected one of: early, mid, end, terminal")

        print(json.dumps({"status": "ok", "output_path": output_path}))

    except Exception as e:
        print(json.dumps({"status": "error", "message": str(e)}))
        sys.exit(1)


if __name__ == "__main__":
    main()
