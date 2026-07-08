"""
Stage 4: Phase-wise CO Attainment Calculation.

Supports:
- early/mid: interim report from internal-only data
- end: final report from internal + terminal (end-sem) data

TERMINAL FILE — NEW CAMU FORMAT:
  Row 1 (openpyxl): Question number / code labels              (cols G+)
  Row 2 (openpyxl): "Ques No X - COY" header with CO tag      (cols G+)
  Row 3 (openpyxl): Max marks per question                     (cols G+)
  Row 4 (openpyxl): (blank row)
  Row 5+           : Student data  Col A=RegNo, Col B=Name, Col G+=marks
"""

import json
import re
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from utils import validate_file_exists


# ---------------------------------------------------------------------------
# Terminal CAMU file layout (1-indexed, openpyxl)
# ---------------------------------------------------------------------------
TERM_ROW_Q_LABELS   = 1   # Row 1 : question number codes
TERM_ROW_CO_HEADERS = 2   # Row 2 : "Ques No X - COY"  headers
TERM_ROW_MAX_MARKS  = 3   # Row 3 : max marks per question column
TERM_STUDENT_START  = 5   # Row 5 : first student row
TERM_MARKS_START_COL = 7  # Col G (1-indexed): first question marks column

CO_HEADER_PATTERN = re.compile(r"CO\d+", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def rewrite_formula(formula):
    """
    Rewrite Excel formula to add Sheet1! prefix to cell references.

    When copying formulas from Sheet1 to Sheet2, all cell references must
    be prefixed with "Sheet1!" to still point to Sheet1 data.

    Example: =A1+B2 -> =Sheet1!A1+Sheet1!B2
    """
    if formula is None or not str(formula).startswith("="):
        return formula

    formula_str = str(formula)
    content = formula_str[1:]

    def add_sheet_prefix(match):
        return f"Sheet1!{match.group(0)}"

    rewritten = re.sub(r"(\$?[A-Z]{1,3}\$?\d+)", add_sheet_prefix, content)
    return "=" + rewritten


def parse_float(value, default=0.0):
    try:
        if value is None:
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def extract_student_rows(sheet, start_row=8, regno_col=2):
    """Return continuous student rows based on REGNO column."""
    rows = []
    row = start_row
    max_guard = max(sheet.max_row + 5, start_row)

    while row <= max_guard:
        regno = sheet.cell(row=row, column=regno_col).value
        if regno is None or str(regno).strip() == "":
            break
        rows.append(row)
        row += 1

    return rows


def create_or_reset_sheet2(wb):
    if "Sheet2" in wb.sheetnames:
        sheet2 = wb["Sheet2"]
        if sheet2.max_row > 0:
            sheet2.delete_rows(1, sheet2.max_row)
    else:
        sheet2 = wb.create_sheet("Sheet2")
    return sheet2


def generate_feedback_sheet(wb, sheet2, student_rows, ep, phase, internal_start_col=5, co_count=6):
    """
    Create a Feedback sheet for early intervention.

    Includes students whose average internal CO score is below EP.
    """
    if phase not in {"early", "mid"}:
        return

    if "Feedback" in wb.sheetnames:
        feedback_ws = wb["Feedback"]
        if feedback_ws.max_row > 0:
            feedback_ws.delete_rows(1, feedback_ws.max_row)
    else:
        feedback_ws = wb.create_sheet("Feedback")

    feedback_ws.cell(row=1, column=1).value = "S.No"
    feedback_ws.cell(row=1, column=2).value = "REGNO"
    feedback_ws.cell(row=1, column=3).value = "Name"

    output_row = 2
    for row in student_rows:
        internal_values = []
        for i in range(co_count):
            val = sheet2.cell(row=row, column=internal_start_col + i).value
            parsed = parse_float(val, default=None)
            if parsed is not None:
                internal_values.append(parsed)

        if not internal_values:
            continue

        internal_score = sum(internal_values) / len(internal_values)
        if internal_score < ep:
            feedback_ws.cell(row=output_row, column=1).value = sheet2.cell(row=row, column=1).value
            feedback_ws.cell(row=output_row, column=2).value = sheet2.cell(row=row, column=2).value
            feedback_ws.cell(row=output_row, column=3).value = sheet2.cell(row=row, column=3).value
            output_row += 1

def write_summary_block(sheet, summary_row, label_col, start_col, co_count, last_student, ep, ela_values):
    sheet.cell(row=summary_row, column=label_col).value = "Total"
    for i in range(co_count):
        col = get_column_letter(start_col + i)
        sheet.cell(row=summary_row, column=start_col + i).value = (
            f"=COUNT({col}6:{col}{last_student})"
        )

    sheet.cell(row=summary_row + 1, column=label_col).value = "EP"
    for i in range(co_count):
        sheet.cell(row=summary_row + 1, column=start_col + i).value = ep

    sheet.cell(row=summary_row + 2, column=label_col).value = ">=EP"
    for i in range(co_count):
        col = get_column_letter(start_col + i)
        ep_col = get_column_letter(start_col + i)
        sheet.cell(row=summary_row + 2, column=start_col + i).value = (
            f'=COUNTIF({col}6:{col}{last_student},">="&{ep_col}{summary_row + 1})'
        )

    sheet.cell(row=summary_row + 3, column=label_col).value = "Actual Attainment"
    for i in range(co_count):
        c_cell = f"{get_column_letter(start_col + i)}{summary_row + 2}"
        t_cell = f"{get_column_letter(start_col + i)}{summary_row}"
        cell = sheet.cell(row=summary_row + 3, column=start_col + i)
        cell.value = f'=IFERROR(({c_cell}/{t_cell})*100, 0)'
        cell.number_format = "0.00"
        
    ela_row = summary_row + 5
    rel_row = ela_row + 1

    sheet.cell(row=ela_row, column=label_col).value = "ELA"
    sheet.cell(row=rel_row, column=label_col).value = "Relative Attainment (%)"

    for i in range(co_count):
        sheet.cell(row=ela_row, column=start_col + i).value = ela_values[i]
        actual_cell = f"{get_column_letter(start_col + i)}{summary_row + 3}"
        ela_cell = f"{get_column_letter(start_col + i)}{ela_row}"
        
        cell = sheet.cell(row=rel_row, column=start_col + i)
        cell.value = f'=IFERROR(MIN(({actual_cell}/{ela_cell})*100, 100), 0)'
        cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# New CAMU Terminal Parsing
# ---------------------------------------------------------------------------

def _extract_co_from_header(header_value):
    """
    Extract CO tag from a header like 'Ques No 1 - CO2' or 'CO3'.
    Returns the CO tag string (e.g. 'CO2') or None.
    """
    if header_value is None:
        return None
    s = str(header_value).strip()
    match = CO_HEADER_PATTERN.search(s)
    return match.group(0).upper() if match else None


def parse_terminal_camu(term_ws):
    """
    Parse the new CAMU terminal marks Excel worksheet.
    """
    marks_start_col = TERM_MARKS_START_COL

    # Find student start row dynamically
    student_start = 5
    for r in range(3, term_ws.max_row + 1):
        val = term_ws.cell(row=r, column=1).value
        if val is not None and re.match(r"^\d+", str(val).strip()):
            student_start = r
            break

    # Dynamically find max marks row and CO headers row by scanning up from student_start
    row_max_marks = 3
    for r in range(student_start - 1, 0, -1):
        val = term_ws.cell(row=r, column=marks_start_col).value
        if val is not None and parse_float(val, -1.0) >= 0:
            row_max_marks = r
            break

    row_co_headers = 2
    for r in range(row_max_marks - 1, 0, -1):
        val = term_ws.cell(row=r, column=marks_start_col).value
        if _extract_co_from_header(val):
            row_co_headers = r
            break

    # ---- Step 1: Read per-column CO tag and max mark ----
    col_co  = []
    col_max = []

    col = marks_start_col
    while col <= term_ws.max_column:
        header_val = term_ws.cell(row=row_co_headers, column=col).value
        max_val    = term_ws.cell(row=row_max_marks,  column=col).value

        co_tag = _extract_co_from_header(header_val)

        # Stop at first completely blank column
        if header_val is None and max_val is None:
            break

        col_co.append(co_tag)
        col_max.append(parse_float(max_val, 0.0))
        col += 1

    num_q_cols = len(col_co)
    if num_q_cols == 0:
        raise ValueError(
            "No question columns found in terminal CAMU file. "
            "Expected 'Ques No X - COY' headers in row 2 from column G onwards."
        )

    # ---- Step 2: Read student marks (rows 5+), match by RegNo ----
    student_co_pct = {}   # RegNo → {CO: pct}
    student_order  = []   # preserve insertion order for Sheet2 row mapping
    ordered_cos    = []   # COs in order of first appearance

    for co in col_co:
        if co and co not in ordered_cos:
            ordered_cos.append(co)

    for row in range(student_start, term_ws.max_row + 1):
        reg_val = term_ws.cell(row=row, column=1).value
        reg_str = str(reg_val).strip() if reg_val is not None else ""
        if not reg_str or reg_str.lower() in ("nan", "none"):
            break
        if "instruction" in reg_str.lower():
            continue

        # Accumulate: {CO: (sum_obtained, sum_max)}
        co_accum = {}
        for q_idx in range(num_q_cols):
            co_tag = col_co[q_idx]
            if not co_tag:
                continue   # skip untagged columns
            mark_val = term_ws.cell(row=row, column=marks_start_col + q_idx).value
            mark     = parse_float(mark_val, 0.0)
            max_m    = col_max[q_idx]

            if co_tag not in co_accum:
                co_accum[co_tag] = [0.0, 0.0]   # [obtained, max]
            co_accum[co_tag][0] += mark
            co_accum[co_tag][1] += max_m

        # Convert to percentage
        co_pct = {}
        for co_tag, (obtained, total_max) in co_accum.items():
            if total_max > 0:
                co_pct[co_tag] = (obtained / total_max) * 100.0
            else:
                co_pct[co_tag] = 0.0

        student_co_pct[reg_str] = co_pct
        student_order.append(reg_str)

    return col_co, col_max, student_co_pct, ordered_cos, student_order


def get_terminal_co_value(student_co_pct, regno, co_name):
    """
    Look up a student's CO percentage from the terminal data.
    Returns 0.0 if student or CO not found.
    """
    # Try exact match first
    if regno in student_co_pct:
        return student_co_pct[regno].get(co_name, 0.0)
    # Try prefix match (terminal file may have truncated IDs)
    for key in student_co_pct:
        if regno.startswith(key) or key.startswith(regno):
            return student_co_pct[key].get(co_name, 0.0)
    return 0.0


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    try:
        if len(sys.argv) < 2:
            raise ValueError("Missing JSON argument payload")

        args = json.loads(sys.argv[1])
        phase = str(args.get("phase", "end")).strip().lower()

        co_attainment_path = args.get("co_attainment_path")
        terminal_path      = args.get("terminal_path")
        output_path        = args.get("output_path")
        ep                 = parse_float(args.get("ep"), 80)
        constraint         = parse_float(args.get("constraint"), 79.99)
        ela                = args.get("ela", {})

        if phase not in {"early", "mid", "end"}:
            raise ValueError("Invalid phase. Expected one of: early, mid, end")

        if not co_attainment_path:
            raise ValueError("Missing required path: co_attainment_path")
        if not output_path:
            raise ValueError("Missing required path: output_path")

        validate_file_exists(co_attainment_path)
        if phase == "end":
            if not terminal_path:
                raise ValueError("Missing required path for end phase: terminal_path")
            validate_file_exists(terminal_path)

        wb     = load_workbook(co_attainment_path, data_only=False)
        sheet1 = wb["Sheet1"]
        sheet2 = create_or_reset_sheet2(wb)

        # Parse terminal file once if end phase
        term_ws          = None
        student_co_pct   = {}
        ordered_cos_term = []
        if phase == "end":
            term_wb = load_workbook(terminal_path, data_only=True)
            term_ws = term_wb.active
            _, _, student_co_pct, ordered_cos_term, _ = parse_terminal_camu(term_ws)

        start_row    = 8
        src_start    = 80   # Column CB — internal CO % columns in Sheet1
        dst_start    = 5    # Column E  — internal CO % in Sheet2
        co_count     = 6
        student_rows = extract_student_rows(sheet1, start_row=start_row, regno_col=2)

        if not student_rows:
            raise ValueError("No student rows found in Sheet1")

        last_student = student_rows[-1] - 2  # Adjusted for Sheet2

        # ---- Step A: Copy identity columns (A, B, C) ----
        sheet2["A3"] = "S.No"
        sheet2["B3"] = "REGNO"
        sheet2["C3"] = "Student Name"

        for row in student_rows:
            dest_row = row - 2
            sheet2.cell(row=dest_row, column=1).value = sheet1.cell(row=row, column=1).value
            sheet2.cell(row=dest_row, column=2).value = sheet1.cell(row=row, column=2).value
            sheet2.cell(row=dest_row, column=3).value = sheet1.cell(row=row, column=3).value

            sheet2.cell(row=dest_row, column=1).alignment = Alignment(horizontal="center")
            sheet2.cell(row=dest_row, column=2).alignment = Alignment(horizontal="center")
            sheet2.cell(row=dest_row, column=3).alignment = Alignment(horizontal="left")

        # ---- Step B: Internal % section (columns E–J) ----
        sheet2.merge_cells("E2:J2")
        sheet2["E2"] = "CO BASED Percentage of marks (Internal Assessment only)"
        sheet2["E2"].alignment = Alignment(horizontal="center", vertical="center")

        for i in range(co_count):
            sheet2.cell(row=3, column=5 + i).value = f"CO{i + 1}"

        for row in student_rows:
            dest_row = row - 2
            for i in range(co_count):
                src = sheet1.cell(row=row, column=src_start + i)
                dst = sheet2.cell(row=dest_row, column=dst_start + i)

                if src.data_type == "f":
                    dst.value = rewrite_formula(src.value)
                else:
                    dst.value = src.value

                dst.number_format = "0.00"
                dst.alignment = Alignment(horizontal="center")

        # ---- Step C: Terminal section (columns L–Q) ----
        sheet2.merge_cells("L2:Q2")
        if phase == "end":
            sheet2["L2"] = "Terminal Assessment"
        else:
            sheet2["L2"] = "Terminal Assessment (Not available in this phase)"
        sheet2["L2"].alignment = Alignment(horizontal="center", vertical="center")

        for i in range(co_count):
            co_name = f"CO{i + 1}"
            if phase == "end" and ordered_cos_term:
                # Use actual CO name from terminal file if available
                header = co_name if co_name in ordered_cos_term else co_name
            else:
                header = co_name
            sheet2.cell(row=3, column=12 + i).value = header

        if phase == "end":
            # Write per-student CO percentages from the new CAMU terminal format
            for row in student_rows:
                dest_row = row - 2
                # Get the student's RegNo from Sheet1 (col B = column 2)
                regno = str(sheet1.cell(row=row, column=2).value or "").strip()
                for i in range(co_count):
                    co_name = f"CO{i + 1}"
                    pct_val = get_terminal_co_value(student_co_pct, regno, co_name)
                    cell = sheet2.cell(row=dest_row, column=12 + i)
                    cell.value = round(pct_val, 4)
                    cell.number_format = "0.00"
                    cell.alignment = Alignment(horizontal="center")
        else:
            for row in student_rows:
                dest_row = row - 2
                for i in range(co_count):
                    sheet2.cell(row=dest_row, column=12 + i).value = 0

        # ---- Step D: Final CO headers (columns S–X) ----
        for i in range(co_count):
            sheet2.cell(row=3, column=19 + i).value = f"CO{i + 1}"

        # ---- Step E: Phase-aware final CO formula ----
        for row in student_rows:
            dest_row = row - 2
            for i in range(co_count):
                e_col    = get_column_letter(5 + i)
                l_col    = get_column_letter(12 + i)
                dest_col = 19 + i

                cell = sheet2.cell(row=dest_row, column=dest_col)
                if phase == "end":
                    # 60% internal + 40% terminal
                    cell.value = f"=0.6*{e_col}{dest_row}+0.4*{l_col}{dest_row}"
                else:
                    # Interim phases: internal-only attainment
                    cell.value = f"={e_col}{dest_row}"
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="center")

        # ---- Step F: Attainment summary rows ----
        summary_row = last_student + 3
        ela_values = [parse_float(ela.get(f"CO{i + 1}"), 0) for i in range(co_count)]

        # Final CO Attainment summary block
        write_summary_block(sheet2, summary_row, 18, 19, co_count, last_student, ep, ela_values)

        # ---- Step H: Early intervention sheet for interim phases ----
        generate_feedback_sheet(
            wb=wb,
            sheet2=sheet2,
            student_rows=student_rows,
            ep=ep,
            phase=phase,
            internal_start_col=5,
            co_count=co_count,
        )
        if phase == "end" and "Feedback" in wb.sheetnames:
            del wb["Feedback"]

        wb.save(output_path)
        print(json.dumps({"status": "ok", "output_path": output_path}))

    except Exception as exc:
        print(json.dumps({"status": "error", "message": str(exc)}))
        sys.exit(1)


if __name__ == "__main__":
    main()
