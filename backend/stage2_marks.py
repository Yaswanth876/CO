"""
Stage 2: CAMU Marks Parser → QP Excel Builder (New Format)

NEW CAMU FORMAT (both internal CAT and assignment marks):
  Row 1: Question number / code labels  (cols G+)
  Row 2: "Ques No X - " header text
  Row 3: CO tag per question             (cols G+)  e.g. CO1, CO2, CO3
  Row 4: Max marks per question           (cols G+)  e.g. 3, 3, 6, 18
  Row 5+: Student data (Col A = RegNo, Col B = Name)

This script automatically detects the student start row dynamically to support
both standard 5-row header exports and 4-row header exports.
"""

import sys
import json
import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from utils import validate_file_exists

# Output Excel row positions (Stage-3 compatible)
OUT_ROW_Q_LABELS  = 10
OUT_ROW_MAX_MARKS = 11
OUT_ROW_CO_TAGS   = 12
OUT_STUDENT_START = 13
OUT_DATA_START_COL = 4   # Column D (1-indexed): first question column


def _to_numeric(value):
    """Convert a cell value to int/float or return None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip().replace("'", "")
    if not s or s.lower() in ("nan", "none", ""):
        return None
    try:
        if re.fullmatch(r"-?\d+", s):
            return int(s)
        if re.fullmatch(r"-?\d*\.\d+", s):
            return float(s)
    except (TypeError, ValueError):
        pass
    return None


def _clean_co_tag(value):
    """
    Return the CO tag string (e.g. 'CO1') from a cell value,
    or an empty string if not a valid CO tag.
    """
    if value is None:
        return ""
    s = str(value).strip()
    match = re.search(r"CO\d+", s, re.IGNORECASE)
    if match:
        return match.group(0).upper()
    return ""


def map_camu_question_to_template_id(q_label, max_mark=None):
    """
    Map CAMU question labels (e.g. '1', '7.1') to template question IDs (e.g. 'A1', 'C1').
    """
    q_str = str(q_label).strip()

    # If it's already in the format of A1, B1, C1 etc., return it
    if re.match(r"^[A-Za-z]\d+", q_str):
        return q_str.upper()

    # Check if it has a decimal point (choice question like 7.1, 7.2)
    if '.' in q_str:
        parts = q_str.split('.')
        try:
            main_q = int(parts[0])
            sub_q = int(parts[1])
            # Map main_q 7 to C1/C2, 8 to C3/C4, 9 to C5/C6, 10 to C7/C8
            # E.g. 7.1 -> C1, 7.2 -> C2
            # E.g. 8.1 -> C3, 8.2 -> C4
            if main_q >= 7:
                c_idx = (main_q - 7) * 2 + sub_q
                return f"C{c_idx}"
        except ValueError:
            pass

    # If it is a simple integer, map 1-6 to A1-A6, 7+ to C1-C8 (fallback)
    try:
        val = int(float(q_str))
        if val <= 6:
            return f"A{val}"
        else:
            c_idx = (val - 7) * 2 + 1
            return f"C{c_idx}"
    except ValueError:
        pass

    return q_str


def parse_camu_marks(marks_path):
    """
    Parse the new CAMU marks Excel file and return extracted data.
    """
    wb = load_workbook(marks_path, data_only=True)
    ws = wb.active

    # ---- Dynamic Header and Student Row detection ----
    # Find first row where column A has a valid Register Number (digits)
    student_start = 5  # default fallback
    for r in range(3, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val is not None and re.match(r"^\d+", str(val).strip()):
            student_start = r
            break

    # Determine layout dynamically based on student start row
    camu_row_q_labels  = 1
    camu_row_max_marks = student_start - 1
    camu_row_co_tags   = student_start - 2
    camu_student_start = student_start
    camu_marks_start_col = 7  # Column G (1-indexed)

    # ---- Read question metadata (cols G+ across rows) ----
    q_labels  = []
    co_tags   = []
    max_marks = []

    col = camu_marks_start_col
    while col <= ws.max_column:
        q_label  = ws.cell(row=camu_row_q_labels,  column=col).value
        co_raw   = ws.cell(row=camu_row_co_tags,   column=col).value
        max_raw  = ws.cell(row=camu_row_max_marks,  column=col).value

        # Stop when headers are blank
        if q_label is None and co_raw is None and max_raw is None:
            break

        co_str = _clean_co_tag(co_raw)

        # Question label: map to template ID
        if q_label is not None:
            label = map_camu_question_to_template_id(q_label, max_raw)
        else:
            label = f"A{col - camu_marks_start_col + 1}"

        q_labels.append(label)
        co_tags.append(co_str)
        max_marks.append(_to_numeric(max_raw) or 0)
        col += 1

    num_q_cols = len(q_labels)
    if num_q_cols == 0:
        raise ValueError(
            "No question columns found in CAMU marks file. "
            "Expected questions starting from column G."
        )

    # ---- Read student data (rows student_start+) ----
    regnos       = []
    names        = []
    marks_matrix = []

    for row in range(camu_student_start, ws.max_row + 1):
        reg_val  = ws.cell(row=row, column=1).value   # Col A
        name_val = ws.cell(row=row, column=2).value   # Col B

        # Stop at blank/invalid rows
        reg_str = str(reg_val).strip() if reg_val is not None else ""
        if not reg_str or reg_str.lower() in ("nan", "none"):
            break
        if "instruction" in reg_str.lower():
            continue

        regnos.append(reg_str)
        names.append(str(name_val).strip() if name_val not in (None, "") else "")

        # Read per-question marks
        row_marks = []
        for q_idx in range(num_q_cols):
            cell_val = ws.cell(row=row, column=camu_marks_start_col + q_idx).value
            row_marks.append(_to_numeric(cell_val))   # None = absent/blank
        marks_matrix.append(row_marks)

    if not regnos:
        raise ValueError(
            f"No student rows found in CAMU marks file starting at row {camu_student_start}."
        )

    return {
        "q_labels":    q_labels,
        "co_tags":     co_tags,
        "max_marks":   max_marks,
        "regnos":      regnos,
        "names":       names,
        "marks_matrix": marks_matrix,
    }


def build_output_excel(output_path, q_labels, co_tags, max_marks, regnos, names, marks_matrix):
    """
    Write a Stage-3-compatible Excel file from the parsed CAMU data.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ---- Write question header rows ----
    for j, (label, max_m, co) in enumerate(zip(q_labels, max_marks, co_tags)):
        data_col = OUT_DATA_START_COL + j
        ws.cell(row=OUT_ROW_Q_LABELS,  column=data_col).value = label
        ws.cell(row=OUT_ROW_MAX_MARKS, column=data_col).value = max_m
        ws.cell(row=OUT_ROW_CO_TAGS,   column=data_col).value = co

    # ---- Write student rows ----
    for i, (reg, name, marks) in enumerate(zip(regnos, names, marks_matrix)):
        row_num = OUT_STUDENT_START + i
        ws.cell(row=row_num, column=1).value = i + 1   # Serial
        ws.cell(row=row_num, column=2).value = reg      # RegNo
        ws.cell(row=row_num, column=3).value = name     # Name

        for j, mark in enumerate(marks):
            ws.cell(row=row_num, column=OUT_DATA_START_COL + j).value = mark

    # ---- Auto-size columns for readability ----
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 8)

    wb.save(output_path)


def process_marks(marks_path, output_path):
    """
    High-level entry point: parse CAMU marks file → write Stage-3 Excel.
    """
    data = parse_camu_marks(marks_path)
    build_output_excel(
        output_path,
        data["q_labels"],
        data["co_tags"],
        data["max_marks"],
        data["regnos"],
        data["names"],
        data["marks_matrix"],
    )
    return data


def main():
    try:
        if len(sys.argv) < 2:
            raise ValueError("Missing JSON argument payload")

        args = json.loads(sys.argv[1])
        marks_path  = args.get("student_db_path") or args.get("marks_path")
        output_path = args.get("output_path")

        if not marks_path:
            raise ValueError("Missing required argument: student_db_path")
        if not output_path:
            raise ValueError("Missing required argument: output_path")

        validate_file_exists(marks_path)

        process_marks(marks_path, output_path)

        print(json.dumps({"status": "ok", "output_path": output_path}))

    except Exception as e:
        print(json.dumps({"status": "error", "message": str(e)}))
        sys.exit(1)


if __name__ == "__main__":
    main()
